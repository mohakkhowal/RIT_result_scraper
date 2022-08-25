#imports
import webbrowser
from bs4 import BeautifulSoup as bs
import requests
from alive_progress import alive_bar
import openpyxl as op
import sys
from openpyxl.styles import Font
#i/p o/p file details: 
usnList = input("Enter name of desired list of USN, (in sample.txt-like format): ")
resultFile = input("Enter name of desired Result .xlsx file(without extension), and PLEASE CLOSE THE FILE, IF ALREADY OPENED: ")+".xlsx"

try:
    file = open(usnList,'r')
except:
    print("Incorrect file name")
    sys.exit(0)

wb = op.Workbook()
allUsn = file.read().split()
total = len(allUsn)
faildata = {}
count=1
with alive_bar(total) as bar:
    for usn in allUsn:
        bar()
        URL = "http://exam.msrit.edu/index.php/component/examresult/?usn="+usn+"&task=getResult&stage=1"
        branch = usn[5:7]
        try:
            has_f_x_i=False
            r = requests.get(URL)
            soup = bs(r.text, "lxml")
            name = soup.find("h3").contents[0]
            # imgUrl = "http://exam.msrit.edu"+soup.find(class_="uk-preserve-width uk-border").attrs['src']
            # # print(imgUrl)
            # response = requests.get(imgUrl, stream=True)
            # with open('img.png', 'wb') as out_file:
            #     shutil.copyfileobj(response.raw, out_file)
            # del response
            basicdata = soup.find(class_= "detail3").find_all("p")
            cred_reg = basicdata[0].contents[0]
            cred_earn = basicdata[1].contents[0]
            sgpa = basicdata[2].contents[0]
            cgpa = basicdata[3].contents[0]
            advdata = soup.find(class_="uk-table uk-table-striped res-table").findAll("td")
            i=0
            toinsert = [usn,name,cred_reg,cred_earn,sgpa,cgpa]
            while(i<len(advdata)):
                toinsert.append(''.join([x[0] for x in advdata[i+1].contents[0].split()])+'/'+advdata[i].contents[0])
                toinsert.append(advdata[i+4].contents[0])
                if((advdata[i+4].contents[0]=='X')or(advdata[i+4].contents[0]=='I')or(advdata[i+4].contents[0]=='F')):
                    has_f_x_i=True
                    try:
                        faildata[''.join([x[0] for x in advdata[i+1].contents[0].split()])+'/'+advdata[i].contents[0]]+=1
                    except:
                        faildata[''.join([x[0] for x in advdata[i+1].contents[0].split()])+'/'+advdata[i].contents[0]]=1
                i+=5
            sheet = wb.active
            # count = 2

            if(branch not in wb.sheetnames):
                if(count-1):
                    # print(faildata,count)
                    sheet.append([])
                    sheet.append([])
                    sheet.append(["Failure Data: "])
                    for a,b in faildata.items():
                        sheet.append([a,b])

                faildata = {}
                wb.create_sheet(branch)
                wb.active = wb[branch]
                sheet = wb.active
                count=1
                sheet.append(["USN","Name","Creds Registered","Creds Earned","SGPA","CGPA","Subject","Grade"])
            # img = op.drawing.image.Image('img.png')
            # img.height = 100
            # img.width=100
            # sheet.row_dimensions[count].height = 100
            # sheet.column_dimensions['A'].width = 30
            # img.anchor = 'A'+str(count)
            count+=1
            # sheet.add_image(img)
            sheet.append(toinsert)
            # count+=1
            # print(usn+"Done")
            if(has_f_x_i):
                try:
                    faildata["Subs >=1"]+=1
                except:
                    faildata["Subs >=1"]=1
                for cell in sheet[str(count)+":"+str(count)]:
                    cell.font = Font(color='00FF0000', italic=True)
            # wb.save(filename=resultFile)
        except:
            print("Unexpected Error :( at USN: ",usn," please check if it's a valid USN, or still in database")
wb.remove(wb['Sheet'])
# print(faildata,count)
if(count-1):
    # print(faildata,count)
    sheet.append([])
    sheet.append([])
    sheet.append(["Failure Data: "])
    for a,b in faildata.items():
        sheet.append([a,b])
wb.save(filename=resultFile)
print("Generated Result file")
try:
    print("Attempting to open")
    webbrowser.open(resultFile)
except:
    print("Failed to open")
    sys.exit()
sys.exit()